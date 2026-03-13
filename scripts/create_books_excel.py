#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Kutubkhana - Create Excel from book images extraction
Extracts book info from Kutubkhana-pics and creates Excel with library columns.
Columns match: اسم الكتاب, المؤلف, القسم, المحقق, الأجزاء, دار النشر, السنة, النسخ, الحالة, الصندوق, الطاق, ملاحظات
"""

import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

# Book data extracted from Kutubkhana-pics images (language preserved as in source)
BOOKS = [
    {
        "name": "কোরআন পড়ি কোরআন মানি",
        "author": "অধ্যাপক মোহাম্মদ ইমাদুল হক",
        "category": "তাফসীর/কুরআন",
        "editor": "",
        "parts": 1,
        "publisher": "দারুস-সালাম (অধ্যাপিকা ডাঃ আজিজা বেগম)",
        "year": "২০০২",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "প্রথম প্রকাশ: ফেব্রুয়ারী ২০০২",
    },
    {
        "name": "হুকুকুল কুরআন",
        "author": "ফকীহুল আস্ত্র হযরত মাওলানা মুফতী রশীদ আহমাদ ছাহেব লুধিয়ানভী রহ.",
        "category": "তাফসীর/কুরআন",
        "editor": "মাওলানা মুহাম্মাদ আবদুল্লাহ মা'রূফ (অনুবাদক)",
        "parts": 1,
        "publisher": "মাকতাবাতুল আশরাফ",
        "year": "২০২১",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "জানুয়ারি ২০২১, ISBN: 978-984-91730-2-1",
    },
    {
        "name": "জীবন সাজানোর গল্প",
        "author": "আরিফুল ইসলাম",
        "category": "তাযকিয়া/সাধারণ",
        "editor": "ইলিয়াস হাসান",
        "parts": 1,
        "publisher": "আশরাফিয়া বুক হাউজ",
        "year": "২০১৯",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "চতুর্থ মুদ্রণ জুলাই ২০১৯",
    },
    {
        "name": "মাআরিফুল কুরআনের গল্প-ঘটনা",
        "author": "মুফতী মুহাম্মাদ শফী (রহ.)",
        "category": "তাফসীর/কুরআন",
        "editor": "",
        "parts": 1,
        "publisher": "মাকতাবাতুল হাসান",
        "year": "২০১৮",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "অক্টোবর ২০১৮, ISBN: 978-984-8012-10-9",
    },
    {
        "name": "কুরআনের অমর কাহিনীগুচ্ছ - আসহাবুল কুরআন",
        "author": "নাসীম আরাফাত",
        "category": "তাফসীর/কুরআন",
        "editor": "",
        "parts": 1,
        "publisher": "মাকতাবাতুল আখতার",
        "year": "২০১০",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "প্রথম প্রকাশ ফেব্রুয়ারি ২০১০, ISBN: 978-984-8807-08-8",
    },
    {
        "name": "নির্বাচিত হাদীস শরীফ",
        "author": "শাইখুল ইসলাম মুফতী মুহাম্মাদ তাকী উসমানী",
        "category": "হাদীস",
        "editor": "মাওলানা মুহাম্মাদ আবদুল্লাহ মা'রূফ (অনুবাদ ও সম্পাদনা)",
        "parts": 1,
        "publisher": "মাকতাবাতুল আশরাফ",
        "year": "২০১৮",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "আগস্ট ২০১৮, ISBN: 978-984-91725-7-4",
    },
    {
        "name": "হাদীস রত্ন ভাণ্ডার বা শাসন পদ্ধতি",
        "author": "মুজাহিদে আযম হযরত মাওলানা শামছুল হক ফরিদপুরী ছদর সাহেব হুযুর (রহ.)",
        "category": "হাদীস",
        "editor": "",
        "parts": 1,
        "publisher": "রশিদিয়া লাইব্রেরী (পরিবেশক)",
        "year": "২০০৭",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "ষষ্ঠ মুদ্রণ জুন ২০০৭",
    },
    {
        "name": "হাদীসের দর্পণে আমাদের কাল",
        "author": "শহীদে খতমে নবুওয়ত হযরত মাওলানা ইউসুফ লুধয়ানবী (রহ.)",
        "category": "হাদীস",
        "editor": "মুহাম্মদ যাইনুল আবিদীন (অনুবাদক)",
        "parts": 1,
        "publisher": "মাকতাবাতুল আখতার",
        "year": "২০০৭",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "প্রথম প্রকাশ জুলাই ২০০৭",
    },
    {
        "name": "ছোটদের হাদিসের গল্প",
        "author": "মুনীরুল ইসলাম",
        "category": "হাদীস",
        "editor": "",
        "parts": 1,
        "publisher": "আনোয়ার লাইব্রেরী",
        "year": "২০২১",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "জুলাই ২০২১",
    },
    {
        "name": "কুরআন শিক্ষার সামাজিকীকরণ",
        "author": "আবু সাঈদ খান",
        "category": "তাফসীর/কুরআন",
        "editor": "",
        "parts": 1,
        "publisher": "কুরআন একাডেমি ফাউন্ডেশন (কাফ)",
        "year": "২০২৩",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "মার্চ ২০২৩, ISBN: 978-984-93957-37",
    },
    {
        "name": "ছোটদের আল-কুরআনের মানুষ",
        "author": "ড. ইকবাল কবীর মোহন",
        "category": "তাফসীর/কুরআন",
        "editor": "",
        "parts": 1,
        "publisher": "শিশু কানন",
        "year": "২০২২",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "চতুর্দশ মুদ্রণ অক্টোবর ২০২২, প্রথম প্রকাশ ফেব্রুয়ারি ২০১১",
    },
    {
        "name": "উন্নত জীবন",
        "author": "ড. লুৎফর রহমান",
        "category": "সাধারণ",
        "editor": "",
        "parts": 1,
        "publisher": "আদিত্য অনীক প্রকাশনী",
        "year": "২০১৬",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "অক্টোবর ২০১৬, ISBN: 978-984-92414-7-8",
    },
    # --- Kutubkhana-pics/kutub (Arabic books) ---
    {
        "name": "في المال والاقتصاد والملكية والعقد دراسة فقهية قانونية اقتصادية",
        "author": "أ. د. علي محيي الدين القره داغي",
        "category": "فقه",
        "editor": "",
        "parts": 1,
        "publisher": "دار البشائر الإسلامية",
        "year": "2009",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "الطبعة الثانية، ١٤٣٠ هـ - ٢٠٠٩ م",
    },
    {
        "name": "كتاب الديات",
        "author": "الإمام الحافظ أبي بكر بن أبي عاصم",
        "category": "فقه",
        "editor": "الدكتور عادل حسن علي",
        "parts": 1,
        "publisher": "المختار للنشر والتوزيع",
        "year": "",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "حققه وخرج أحاديثه",
    },
    {
        "name": "النشأة الثانية للفقه الإسلامي (المذهب الحنفي في فجر الدولة العثمانية الحديثة)",
        "author": "جاي بوراك",
        "category": "فقه",
        "editor": "د. أحمد محمود إبراهيم / د. أسامة شفيع السيد (ترجمة)",
        "parts": 1,
        "publisher": "مركز نماء للبحوث والدراسات",
        "year": "2018",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "الطبعة الأولى، ISBN: 978-614-431-719-8",
    },
    {
        "name": "الاختلاف بين الأئمة أسبابه وعوامله",
        "author": "العَلَامَة المحدث الشيخ محمد زكريا الكاندهلوي المدني",
        "category": "فقه",
        "editor": "محمد محمد معاوية سَعْدِي (تحقيق)",
        "parts": 1,
        "publisher": "مكتبة العصري",
        "year": "",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "تعريب: محمد مجد القاسمي",
    },
    {
        "name": "أصول مذاهب الأئمة الأوزاعي من واقع فقهه وآثاره",
        "author": "أ. د. علي بن سعد بن صالح الضويحي",
        "category": "فقه",
        "editor": "",
        "parts": 1,
        "publisher": "الرسالة العالمية",
        "year": "2013",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "الطبعة الأولى، ١٤٣٤ هـ - ٢٠١٣ م",
    },
    {
        "name": "فلسفة ميراث الأنثى في الإسلام",
        "author": "د. أحمد حسن العيساوي",
        "category": "فقه",
        "editor": "",
        "parts": 1,
        "publisher": "دار الرسالة العالمية",
        "year": "2022",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "الطبعة الأولى، ISBN: 978-9933-424-84-8",
    },
    {
        "name": "الغش وأثره في العقود",
        "author": "الدكتور عبد الله بن ناصر السالمي",
        "category": "فقه",
        "editor": "",
        "parts": 1,
        "publisher": "دار كنوز إشبيليا للنشر والتوزيع",
        "year": "2004",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "الجزء الأول، الطبعة الأولى",
    },
    {
        "name": "أحكام التعامل في الأسواق المالية المعاصرة",
        "author": "الدكتور مبارك بن سليمان بن محمد آل سليمان",
        "category": "فقه",
        "editor": "",
        "parts": 1,
        "publisher": "دار كنوز إشبيليا للنشر والتوزيع",
        "year": "2005",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "الجزء الأول، الطبعة الأولى",
    },
    {
        "name": "فقه الهندسة المالية الإسلامية دراسة تأصيلية تطبيقية",
        "author": "د. مرضي بن مسرح العنزي",
        "category": "فقه",
        "editor": "",
        "parts": 1,
        "publisher": "دار كنوز إشبيليا للنشر والتوزيع",
        "year": "2015",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "الطبعة الأولى، ISBN: 978-603-8155-54-7",
    },
    {
        "name": "التدليس في عقد البيع وعقد الزواج وأثره",
        "author": "الدكتور أيمن إبراهيم مَلَض",
        "category": "فقه",
        "editor": "",
        "parts": 1,
        "publisher": "دار الرسالة العالمية",
        "year": "2015",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "الطبعة الأولى، مكتبة دار العلوم",
    },
    {
        "name": "الاقتصاد في التقليد والاجتهاد",
        "author": "العَلَامَة الشيخ حكيم الأمة أشرف علي التهانوي",
        "category": "فقه",
        "editor": "أحمد نقي، ومحمود زكي (ترجمة)",
        "parts": 1,
        "publisher": "مؤسسة المعارف لإحياء التراث الإسلامي",
        "year": "",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "فقه بالأردية، مدينة شيتاغونغ بنغلاديش",
    },
    {
        "name": "مجلة البحوث الفقهية المعاصرة",
        "author": "عدة مؤلفين",
        "category": "فقه",
        "editor": "",
        "parts": 1,
        "publisher": "",
        "year": "2006",
        "copies": 1,
        "status": "متاح",
        "cabinet": "",
        "shelf": "",
        "notes": "مجلة علمية محكمة متخصصة في الفقه الإسلامي",
    },
]

# Arabic headers matching library schema (from data.js)
HEADERS = [
    "اسم الكتاب",
    "المؤلف",
    "القسم",
    "المحقق",
    "الأجزاء",
    "دار النشر",
    "السنة",
    "النسخ",
    "الحالة",
    "الصندوق",
    "الطاق",
    "ملاحظات",
]

# Bengali headers for reference (optional)
HEADERS_BN = [
    "বইয়ের নাম",
    "লেখক",
    "বিভাগ",
    "সম্পাদক/মুহাক্কিক",
    "খণ্ড",
    "প্রকাশক",
    "সাল",
    "কপি",
    "অবস্থা",
    "ক্যাবিনেট",
    "শেলফ",
    "মন্তব্য",
]


def escape_csv_cell(val):
    """Escape a cell for CSV (wrap in quotes, escape internal quotes)."""
    s = "" if val is None else str(val)
    return '"' + s.replace('"', '""') + '"'


def bengali_to_ascii_digits(s):
    """Convert Bengali numerals (০-৯) to ASCII (0-9) so app's year validation accepts them."""
    if not s:
        return s
    tbl = str.maketrans("০১২৩৪৫৬৭৮৯", "0123456789")
    return str(s).translate(tbl)


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.join(script_dir, "..", "Kutubkhana-pics")
    excel_path = os.path.join(out_dir, "books_extracted.xlsx")
    csv_path = os.path.join(out_dir, "books_extracted.csv")
    csv_path_alt = os.path.join(out_dir, "books_extracted_import.csv")

    # Use placeholder for empty cabinet so app import won't skip rows
    cabinet_placeholder = "১"

    wb = Workbook()
    ws = wb.active
    ws.title = "Books"

    # Header row
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows (use cabinet placeholder if empty so app import accepts)
    # Year: convert Bengali numerals to ASCII so app's year validation accepts
    for row_idx, book in enumerate(BOOKS, 2):
        cab = book["cabinet"] if book.get("cabinet") else cabinet_placeholder
        year_val = bengali_to_ascii_digits(book.get("year") or "")
        ws.cell(row=row_idx, column=1, value=book["name"])
        ws.cell(row=row_idx, column=2, value=book["author"])
        ws.cell(row=row_idx, column=3, value=book["category"])
        ws.cell(row=row_idx, column=4, value=book["editor"])
        ws.cell(row=row_idx, column=5, value=book["parts"])
        ws.cell(row=row_idx, column=6, value=book["publisher"])
        ws.cell(row=row_idx, column=7, value=year_val)
        ws.cell(row=row_idx, column=8, value=book["copies"])
        ws.cell(row=row_idx, column=9, value=book["status"])
        ws.cell(row=row_idx, column=10, value=cab)
        ws.cell(row=row_idx, column=11, value=book["shelf"])
        ws.cell(row=row_idx, column=12, value=book["notes"])

    # Adjust column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 6
    ws.column_dimensions["I"].width = 8
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 8
    ws.column_dimensions["L"].width = 25

    wb.save(excel_path)
    print(f"Excel created: {os.path.abspath(excel_path)}")

    # Also create CSV for app import (استيراد من CSV)
    try:
        csv_out = csv_path
        f = open(csv_path, "w", encoding="utf-8-sig")
    except PermissionError:
        csv_out = csv_path_alt
        f = open(csv_path_alt, "w", encoding="utf-8-sig")
    with f:
        f.write(",".join(escape_csv_cell(h) for h in HEADERS) + "\n")
        for book in BOOKS:
            cab = book["cabinet"] if book.get("cabinet") else cabinet_placeholder
            year_val = bengali_to_ascii_digits(book.get("year") or "")
            row = [
                book["name"], book["author"], book["category"], book["editor"],
                book["parts"], book["publisher"], year_val, book["copies"],
                book["status"], cab, book["shelf"], book["notes"]
            ]
            f.write(",".join(escape_csv_cell(v) for v in row) + "\n")
    print(f"CSV created: {os.path.abspath(csv_out)}")
    print(f"Total books: {len(BOOKS)}")
    print("\nFor upload: Use books_extracted.csv in your app (Import from CSV).")


if __name__ == "__main__":
    main()
